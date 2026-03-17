"""
Script to populate S-Core CSV with scores from student REKAPITULASI S-CORE.xlsx files.

Reads student data from public/sources/<NIM - Name>/REKAPITULASI S-CORE.xlsx
and fills in the corresponding scores in the CSV file at public/data/.

Categories mapping (REKAP STATUS sheet → CSV columns):
  - ORKESS           → OrKeSS (WAJIB)
  - RETREAT          → Retreat (WAJIB)          [empty if not found]
  - KEGIATAN ILMIAH  → Kegiatan Ilmiah dan Penalaran
  - ORGANISASI       → Kepengurusan Organisasi/Kepanitiaan
  - PERLOMBAAN       → Performance, Pengembangan, dan Perlombaan
  - SOSIAL           → Kegiatan Sosial Kemasyarakatan

If Retreat is not present → leave empty.
If any other category is not present → fill with 0.
"""

import os
import csv
import openpyxl
import re

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCES_DIR = os.path.join(BASE_DIR, "public", "sources")
CSV_PATH = os.path.join(
    BASE_DIR,
    "public",
    "data",
    "Data email itbss mahasiswa angkatan dan jumlah s core 1-4_program s-core.csv",
)
DELIMITER = ","


def extract_nim_from_folder(folder_name):
    """Extract NIM from folder name like '22100001 - Richard Marcell'."""
    match = re.match(r"^(\d+)", folder_name.strip())
    return match.group(1) if match else None


def read_scores_from_xlsx(xlsx_path):
    """
    Read scores from REKAPITULASI S-CORE.xlsx 'REKAP STATUS' sheet.
    Returns a dict with category keys mapped to their scores.
    """
    scores = {
        "orkess": 0,
        "retreat": "",  # empty by default (not 0)
        "ilmiah": 0,
        "organisasi": 0,
        "perlombaan": 0,
        "sosial": 0,
    }

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"  [ERROR] Cannot open {xlsx_path}: {e}")
        return scores

    if "REKAP STATUS" not in wb.sheetnames:
        print(f"  [WARNING] No 'REKAP STATUS' sheet in {xlsx_path}")
        return scores

    ws = wb["REKAP STATUS"]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=5, values_only=False):
        cell_a = row[0].value
        cell_c = row[2].value if len(row) > 2 else None

        if cell_a is None:
            continue

        cell_a_upper = str(cell_a).upper().strip()

        # Match categories
        if "ORKESS" in cell_a_upper:
            scores["orkess"] = cell_c if cell_c is not None else 0
        elif "RETREAT" in cell_a_upper:
            scores["retreat"] = cell_c if cell_c is not None else ""
        elif "ILMIAH" in cell_a_upper or "PENALARAN" in cell_a_upper:
            scores["ilmiah"] = cell_c if cell_c is not None else 0
        elif "ORGANISASI" in cell_a_upper or "KEPANITIAAN" in cell_a_upper:
            scores["organisasi"] = cell_c if cell_c is not None else 0
        elif "PERLOMBAAN" in cell_a_upper or "KOMPETISI" in cell_a_upper or "PERFORMANCE" in cell_a_upper:
            scores["perlombaan"] = cell_c if cell_c is not None else 0
        elif "SOSIAL" in cell_a_upper or "KEMASYARAKATAN" in cell_a_upper:
            scores["sosial"] = cell_c if cell_c is not None else 0

    wb.close()
    return scores


def build_student_scores_map():
    """
    Scan public/sources/ for student folders and build a map of NIM → scores.
    """
    nim_scores = {}

    if not os.path.isdir(SOURCES_DIR):
        print(f"[ERROR] Sources directory not found: {SOURCES_DIR}")
        return nim_scores

    for folder_name in os.listdir(SOURCES_DIR):
        folder_path = os.path.join(SOURCES_DIR, folder_name)
        if not os.path.isdir(folder_path):
            continue

        nim = extract_nim_from_folder(folder_name)
        if not nim:
            print(f"  [SKIP] Cannot extract NIM from folder: {folder_name}")
            continue

        xlsx_path = os.path.join(folder_path, "REKAPITULASI S-CORE.xlsx")
        if not os.path.isfile(xlsx_path):
            print(f"  [SKIP] No REKAPITULASI S-CORE.xlsx in {folder_name}")
            continue

        print(f"  [READ] {folder_name}")
        scores = read_scores_from_xlsx(xlsx_path)
        nim_scores[nim] = scores

    return nim_scores


def populate_csv(nim_scores):
    """
    Read the CSV, fill in scores for matching students, and write back.
    """
    # Read existing CSV
    with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
        content = f.read()

    # Parse CSV
    lines = content.strip().split("\n")
    header = lines[0]
    header_fields = header.split(DELIMITER)

    print(f"\n[INFO] CSV has {len(lines) - 1} student rows")
    print(f"[INFO] Found scores for {len(nim_scores)} students from sources\n")

    # Find column indices
    # Columns: Nama;Email;Password;student_id;major;batch_year;OrKeSS (WAJIB);Retreat (WAJIB);
    #          Kegiatan Ilmiah dan Penalaran;Performance, Pengembangan, dan Perlombaan;
    #          Kepengurusan Organisasi/Kepanitiaan;Kegiatan Sosial Kemasyarakatan
    col_map = {}
    for i, h in enumerate(header_fields):
        h_clean = h.strip()
        if "OrKeSS" in h_clean:
            col_map["orkess"] = i
        elif "Retreat" in h_clean:
            col_map["retreat"] = i
        elif "Ilmiah" in h_clean or "Penalaran" in h_clean:
            col_map["ilmiah"] = i
        elif "Performance" in h_clean or "Perlombaan" in h_clean:
            col_map["perlombaan"] = i
        elif "Organisasi" in h_clean or "Kepanitiaan" in h_clean:
            col_map["organisasi"] = i
        elif "Sosial" in h_clean or "Kemasyarakatan" in h_clean:
            col_map["sosial"] = i

    print(f"[INFO] Column mapping: {col_map}")

    updated_count = 0
    new_lines = [header]

    for line in lines[1:]:
        fields = line.split(DELIMITER)

        # student_id is at index 3
        student_id = fields[3].strip() if len(fields) > 3 else ""

        if student_id in nim_scores:
            scores = nim_scores[student_id]

            # Ensure fields list is long enough
            while len(fields) < len(header_fields):
                fields.append("")

            # Fill in scores
            for key, col_idx in col_map.items():
                val = scores.get(key, 0 if key != "retreat" else "")
                fields[col_idx] = str(val) if val != "" else ""

            updated_count += 1
            print(f"  [UPDATED] {fields[0]} (NIM: {student_id}) → "
                  f"OrKeSS={scores['orkess']}, Retreat={scores['retreat']}, "
                  f"Ilmiah={scores['ilmiah']}, Organisasi={scores['organisasi']}, "
                  f"Perlombaan={scores['perlombaan']}, Sosial={scores['sosial']}")

        new_lines.append(DELIMITER.join(fields))

    # Write updated CSV (with fallback if file is locked by another process)
    output_path = CSV_PATH
    try:
        with open(CSV_PATH, "w", encoding="utf-8-sig", newline="") as f:
            f.write("\n".join(new_lines) + "\n")
    except PermissionError:
        # File is locked (e.g. open in editor), write to alternate path
        alt_path = CSV_PATH.replace(".csv", "_updated.csv")
        print(f"\n[WARNING] CSV file is locked. Writing to: {os.path.basename(alt_path)}")
        with open(alt_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write("\n".join(new_lines) + "\n")
        output_path = alt_path

    # Report students from sources that were NOT found in CSV (e.g. dropped out)
    csv_nims = set()
    for line in lines[1:]:
        fields = line.split(DELIMITER)
        if len(fields) > 3:
            csv_nims.add(fields[3].strip())

    skipped = [nim for nim in nim_scores if nim not in csv_nims]
    if skipped:
        print(f"\n[SKIPPED] {len(skipped)} student(s) from sources NOT found in CSV (ignored):")
        for nim in skipped:
            print(f"  - NIM {nim}")

    print(f"\n[DONE] Updated {updated_count} student(s) in CSV")
    return updated_count


if __name__ == "__main__":
    print("=" * 60)
    print("S-Core CSV Populator")
    print("=" * 60)
    print(f"\nSources: {SOURCES_DIR}")
    print(f"CSV:     {CSV_PATH}\n")

    print("[STEP 1] Reading student scores from XLSX files...")
    nim_scores = build_student_scores_map()

    print("\n[STEP 2] Populating CSV...")
    populate_csv(nim_scores)

    print("\n[COMPLETE] You can now import the CSV into your database.")
