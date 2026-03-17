import openpyxl
import os
import re

# Paths
storage_path = r'C:\xampp\htdocs\S-Core\storage\app\public'
target_path = r'C:\xampp\htdocs\S-Core\public\data\Data email itbss mahasiswa angkatan dan jumlah s core 1-4_program s-core.xlsx'

# Load target workbook
target_wb = openpyxl.load_workbook(target_path)
target_ws = target_wb['Worksheet']

# Build NIM-to-row mapping from target (column D = student_id)
nim_row_map = {}
for row in range(2, target_ws.max_row + 1):
    nim_val = target_ws.cell(row=row, column=4).value  # column D
    if nim_val:
        nim_row_map[str(nim_val).strip()] = row
print(f'Target file has {len(nim_row_map)} students')

# Scan student folders in storage
folders = [f for f in os.listdir(storage_path) if os.path.isdir(os.path.join(storage_path, f))]
print(f'Found {len(folders)} student folder(s): {folders}')

def sum_column_d(ws, start_row=6):
    """Sum column D values, separating Retreat items."""
    total = 0
    retreat = 0
    for r in range(start_row, ws.max_row + 1):
        d_val = ws.cell(row=r, column=4).value
        b_val = ws.cell(row=r, column=2).value
        if d_val is not None and isinstance(d_val, (int, float)):
            if b_val and 'retreat' in str(b_val).lower():
                retreat += int(d_val)
            else:
                total += int(d_val)
    return total, retreat

updated = 0
for folder in folders:
    # Extract NIM from folder name (format: 'NIM - Name')
    match = re.match(r'^(\d+)', folder)
    if not match:
        print(f'  SKIP: Cannot extract NIM from "{folder}"')
        continue
    nim = match.group(1)

    # Check if REKAPITULASI file exists
    rekap_path = os.path.join(storage_path, folder, 'REKAPITULASI S-CORE.xlsx')
    if not os.path.exists(rekap_path):
        print(f'  SKIP: No REKAPITULASI file for NIM {nim}')
        continue

    # Read student's REKAPITULASI
    try:
        wb = openpyxl.load_workbook(rekap_path)
    except Exception as e:
        print(f'  ERROR reading {rekap_path}: {e}')
        continue

    orkess_pts = 0
    retreat_pts = 0
    ilmiah_pts = 0
    organisasi_pts = 0
    perlombaan_pts = 0
    sosial_pts = 0

    # ORKESS
    if 'ORKESS' in wb.sheetnames:
        pts, ret = sum_column_d(wb['ORKESS'])
        orkess_pts = pts
        retreat_pts += ret

    # KEGIATAN ILMIAH DAN PENALARAN
    if 'KEGIATAN ILMIAH DAN PENALARAN' in wb.sheetnames:
        pts, ret = sum_column_d(wb['KEGIATAN ILMIAH DAN PENALARAN'])
        ilmiah_pts = pts
        retreat_pts += ret

    # ORGANISASI DAN KEPANITIAAN
    if 'ORGANISASI DAN KEPANITIAAN' in wb.sheetnames:
        pts, ret = sum_column_d(wb['ORGANISASI DAN KEPANITIAAN'])
        organisasi_pts = pts
        retreat_pts += ret

    # PERLOMBAAN-KOMPETISI
    if 'PERLOMBAAN-KOMPETISI' in wb.sheetnames:
        pts, ret = sum_column_d(wb['PERLOMBAAN-KOMPETISI'])
        perlombaan_pts = pts
        retreat_pts += ret

    # SOSIAL KEMASYARAKATAN
    if 'SOSIAL KEMASYARAKATAN' in wb.sheetnames:
        pts, ret = sum_column_d(wb['SOSIAL KEMASYARAKATAN'])
        sosial_pts = pts
        retreat_pts += ret

    total_all = orkess_pts + retreat_pts + ilmiah_pts + organisasi_pts + perlombaan_pts + sosial_pts

    print(f'  NIM {nim} ({folder}):')
    print(f'    OrKeSS={orkess_pts}, Retreat={retreat_pts}, Ilmiah={ilmiah_pts}, Organisasi={organisasi_pts}, Perlombaan={perlombaan_pts}, Sosial={sosial_pts}, TOTAL={total_all}')

    # Find row in target
    if nim in nim_row_map:
        row = nim_row_map[nim]
        target_ws.cell(row=row, column=7, value=orkess_pts)     # G: OrKeSS
        target_ws.cell(row=row, column=8, value=retreat_pts)     # H: Retreat
        target_ws.cell(row=row, column=9, value=ilmiah_pts)      # I: Kegiatan Ilmiah
        target_ws.cell(row=row, column=10, value=perlombaan_pts) # J: Perlombaan
        target_ws.cell(row=row, column=11, value=organisasi_pts) # K: Organisasi
        target_ws.cell(row=row, column=12, value=sosial_pts)     # L: Sosial
        updated += 1
        print(f'    -> Updated row {row} in target file')
    else:
        print(f'    -> NIM {nim} NOT FOUND in target file!')

# Save
target_wb.save(target_path)
print(f'\nDone! Updated {updated} student(s) in target file.')
